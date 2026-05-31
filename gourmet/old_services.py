from sqlalchemy import text, create_engine
from flask import jsonify
from os import environ
from hashlib import sha256
from models.now import now, timedelta, dt
from os import getcwd, path
from urllib.parse import unquote_plus
from json import loads
from models.meses import meses
from pandas import read_sql
from openpyxl import Workbook, load_workbook
from os import remove
from threading import Thread
from models.send_email import Email

em = Email()

class Gourmet():
    # Funcoes Primitivas ===================================================
    def __init__(self):
        self.engine = create_engine(environ['gourmet'])
        self.engine_loja = create_engine(environ['lojas'])

    def cons(self, sql, *args, all:bool=None):
        with self.engine.connect() as conn:
            if len(args) == 1: args = args[0]
            txt = sql % args
            res = conn.execute(text(txt))
            res = res.fetchall()
            if len(res) == 1 and not all: return list(res[0]) # Verifica se há somente um valor e o retorna apenas !
            else:
                # A resposta dos valores internos é uma tupla, e nao pode ser convertida para JSON
                # Por isso convertemos internamente para uma lista
                ls = []
                for item in res: ls.append(list(item))
                return ls

    def query(self, sql, *args, commit=True):
        with self.engine.connect() as conn:
            if len(args) == 1: args = args[0]
            txt = sql % args
            res = conn.execute(text(txt))
            if commit: conn.commit()
            try:
                res = res.fetchall()
                if len(res) == 1: return list(res[0]) # Verifica se há somente um valor e o retorna apenas !
                else:
                    # A resposta dos valores internos é uma tupla, e nao pode ser convertida para JSON
                    # Por isso convertemos internamente para uma lista
                    ls = []
                    for item in res: ls.append(list(item))
                    return ls
            except: return 'Sem retorno da Query!'

    def lcons(self, sql, *args, all:bool=None):
        with self.engine_loja.connect() as conn:
            if len(args) == 1: args = args[0]
            txt = sql % args
            res = conn.execute(text(txt))
            res = res.fetchall()
            if len(res) == 1 and not all: return list(res[0]) # Verifica se há somente um valor e o retorna apenas !
            else:
                # A resposta dos valores internos é uma tupla, e nao pode ser convertida para JSON
                # Por isso convertemos internamente para uma lista
                ls = []
                for item in res: ls.append(list(item))
                return ls

    def lquery(self, sql, *args):
        with self.engine_loja.connect() as conn:
            if len(args) == 1: args = args[0]
            txt = sql % args
            res = conn.execute(text(txt))
            conn.commit()
            try:
                res = res.fetchall()
                if len(res) == 1: return list(res[0]) # Verifica se há somente um valor e o retorna apenas !
                else:
                    # A resposta dos valores internos é uma tupla, e nao pode ser convertida para JSON
                    # Por isso convertemos internamente para uma lista
                    ls = []
                    for item in res: ls.append(list(item))
                    return ls
            except: return 'Sem retorno da Query!'

    def get_fuso(self, cr):
        res = self.cons("select fuso from config where cr = '%s'" , (cr))
        return res[0]

    # Login e Cadastro ===================================================
    def login(self, dados):
        matricula = dados.get('matricula')
        res = self.cons('select nome, permissao, grupodecliente, cr, hash from funcionarios where matricula = %s', matricula)
        if res:
            hash = sha256(dados.get('password').encode()).hexdigest()
            hash_res = str(res[-1], 'utf-8')
            if hash_res == hash:
                cr = res[3]
                config = self.get_config(cr)
                return jsonify({
                     'matricula': matricula,
                     'nome': res[0].split()[0], #Pega o Nome completo e divide para que retorne somente o primeiro nome
                     'permissao': res[1],
                     'cr': cr,
                     'gc': res[2],
                     'config': config
                 })
            else: return jsonify('Senha Incorreta'), 401
        else: return jsonify('Matricula Incorreta'), 401

    # Caixa ===================================================
    def status_caixa(self, cr):
        res = self.cons("select valor from caixa_ab where cr = '%s'", cr)
        if res: return True
        else: return False

    def conferir_caixa(self, cr):
        if self.status_caixa(cr): 
            res = self.cons("select valor, caixa_abertura, data from caixa_ab where cr = '%s'", cr)
            valor = res[0] if res else None
            abertura = res[1] if res else None
            data = res[2] if res else None

            if valor and abertura and data:
                res = {
                    'debito': 0,
                    'credito': 0,
                    'pix': 0,
                    'dinheiro': 0,
                    'total': 0,
                    'abertura': abertura,
                    'fechamento': valor,
                }
                caixa = self.cons("select debito, credito, pix, dinheiro, desconto from vendas where cr = '%s' and data >= '%s'", (cr, data), all=True)
                for debito, credito, pix, dinheiro, desconto in caixa:
                    # Confere se tem desconto
                    if desconto > 0:
                        pag = {
                            'debito': debito,
                            'credito': credito,
                            'pix': pix,
                            'dinheiro': dinheiro
                        }

                        # Pega o maior valor estipulando
                        maior = max(pag, key=pag.get)
                        pag[maior] -= desconto

                        # Soma os valores
                        debito = pag['debito']
                        credito = pag['credito']
                        pix = pag['pix']
                        dinheiro = pag['dinheiro']

                    # Repassa os valores ao response
                    res['debito'] += debito
                    res['credito'] += credito
                    res['pix'] += pix
                    res['dinheiro'] += dinheiro
                    res['total'] += (debito + credito + pix + dinheiro)

                return jsonify(res), 200
            return jsonify('Problema com os dados do caixa'), 401
        return jsonify(False), 200

    def abrir_caixa(self, hd, dd):
        # Dados via Headers
        mat = hd.get('mat')
        cr = hd.get('cr')
        perm = hd.get('perm')
        # Dados via Body
        valor = dd.get('valor')

        if mat:
            if cr:
                if valor:
                    if perm == 'ADMIN':
                        if not self.status_caixa(cr):
                            self.query("insert into caixa_ab(cr, valor, data, caixa_abertura) values('%s', %s, '%s', %s)", (cr, valor, now(self.get_fuso(cr)), valor))
                            return jsonify('Sucesso'), 200
                        return jsonify('Caixa já Aberto'), 401
                    return jsonify('Voce não tem permissao!'), 401
                else: return jsonify('Valor Obrigatório!'), 400
            else: return jsonify('CR Obrigatório!'), 400
        else: return jsonify('Matricula Obrigatória!'), 400

    def aplicar_caixa(self, bd, hd):
        mat = hd.get('mat')
        cr = hd.get('cr')
        perm = hd.get('perm')
        
        valor = bd.get('valor')

        if valor:
            valor = int(valor)
            if valor > 0:
                if mat:
                    if cr:
                        if self.status_caixa(cr):
                            if perm == 'ADMIN':
                                valor_ab = self.cons("select caixa_abertura from caixa_ab where cr = '%s'", cr)[0]
                                if not valor_ab: valor_ab = 0
                                self.query("insert into caixa_ap(valor, valor_abertura, matricula, data) values(%s, %s, %s, '%s')", (float(valor), valor_ab, mat, now(self.get_fuso(cr))))
                                self.query("update caixa_ab set valor = valor + %s where cr = '%s'", (float(valor), cr))
                                return jsonify('Sucesso'), 200
                            else: return jsonify('Operação não permitida!'), 401
                        else: return jsonify('Caixa fechado!'), 401
                    else: return jsonify('CR Obrigatorio'), 400
                else: return jsonify('Matricula Obrigatoria'), 400
            else: return jsonify('Valor tem que ser maior que zero!'), 400
        else: return jsonify('Valor Obrigatorio'), 400

    def fechar_caixa(self, hd):
        cr = hd.get('cr')
        gc = hd.get('gc')
        perm = hd.get('perm')
        mat = hd.get('mat')
        
        if self.status_caixa(cr):
            if cr:
                if perm == 'ADMIN':
                    res = self.cons("select data, to_char(data, 'DD/MM/YYYY HH24:MI'), valor, caixa_abertura from caixa_ab where cr = '%s'", cr)
                    data, dataatu, fechamento, abertura = res
                    body = {"valor": data, "filter": "dia", "operador": ">=", "flt": True}
                    filename = self.get_rl_vendas(body, hd)[0]
                    filename = filename.get_data(as_text=True).replace('"', '').replace('\n', '')

                    with self.engine.connect() as conn:
                        cons = """select
                            count(distinct to_char(data, 'DD-MM')) as "Dias Trabalhados",
                            count(distinct id) as "Quantidade Vendas",
                            to_char(AVG(valor_pago - desconto), 'R$ FM9G999G999D00') as "Ticket Médio",
                            case 
                                when sum(valor_pago) > 0 then to_char(sum(valor_pago), 'R$ FM9G999G999D00')
                                else to_char(0, 'R$ FM9G999G9990D00') end as "Total Bruto",
                            case 
                                when sum(valor_pago) - sum(desconto) > 0 then to_char(sum(valor_pago) - sum(desconto), 'R$ FM9G999G999D00')
                                else to_char(0, 'R$ FM9G999G9990D00') end as "Total Liquido"
                            from vw_vendas
                            where data >= '%s'
                            and cr = '%s';""" % (data, cr)
                        df_infos = read_sql(cons, conn)


                    title = f"Fechamento de Caixa | {cr}"
                    html = f"""
                        <body style="background: #333; border-radius: 30px; padding: 30px;">
                            <img src="https://api.hubbix.com.br/img/logo.png" style="width: 30vh;">
                            <p style="font-size: 20px; font-weight: 700;">Fechamento de Caixa</p>

                            <strong>Referência: {dataatu}</strong>
                            <br>
                            <strong>Fechado por: {self.get_name_employe(mat)}</strong>
                            <hr>

                            <br>
                            <strong>Info de Vendas</strong>
                            {df_infos.to_html(index=False)}
                        </body>
                    """
                    mail_to = "ghostlagado@gmail.com"
                    Thread(target=em.send, args=(title, html, mail_to, [filename])).start()
                    self.query("insert into caixa_fc(abertura, fechamento, grupodecliente, cr, data, matricula) values(%s, %s, '%s', '%s', '%s', %s)", (abertura, fechamento, gc, cr, now(self.get_fuso(cr)), mat))
                    self.query("delete from caixa_ab where cr = '%s'" , cr)
                    return jsonify('Sucesso'), 200
                return jsonify('Operação não permitida'), 400
            return jsonify('CR Obrigatorio'), 400
        return jsonify('Caixa já fechado'), 401

    def last_value(self, hd):
        cr = hd.get('cr')
        if cr:
            res = self.cons("select fechamento from caixa_fc where cr = '%s' order by data desc limit 1", cr)
            return jsonify(res), 200
        else: return jsonify('Confira as credenciais - Headers'), 400

    # Despesas ===================================================
    def get_despesas(self, cr, gc):
        dsps = []
        res = self.cons("select id, motivo, valor, to_char(data, 'YYYY-MM-DD HH24:MI:SS') from despesas where cr = '%s' and grupodecliente = '%s' order by data desc", (cr, gc), all=True)
        if cr:
            if gc:
                if res:
                    for id, motivo, valor, data in res:
                        dsps.append(
                            {
                                'id':id,
                                'motivo': motivo,
                                'valor': valor,
                                'data': data
                            }
                        )
                    return jsonify(dsps), 200
                else: return jsonify(dsps), 200
            else: return jsonify('Grupo de cliente obrigatorio'), 400
        else: return jsonify('CR obrigatorio'), 400

    def adicionar_despesas(self, bd, hd):
        valor = bd.get('valor')
        motivo = bd.get('motivo')

        perm = hd.get('perm')
        cr = hd.get('cr')
        gc = hd.get('gc')

        if self.status_caixa(cr):
            if perm == 'ADMIN':
                if valor and valor > 0:
                    if motivo:
                        if cr:
                            if gc:
                                self.query("insert into despesas(motivo, valor, cr, grupodecliente, data) values('%s', %s, '%s', '%s', '%s')", (motivo, valor, cr, gc, now(self.get_fuso(cr))))
                                self.query("update caixa_ab set valor = valor - %s where cr = '%s'", (valor, cr))
                                return jsonify('Sucesso'), 200
                            return jsonify('Grupo de Ciente Obrigatorio'), 400
                        return jsonify('CR Obrigatorio'), 400
                    return jsonify('Motivo Obrigatorio'), 400
                return jsonify('Valor Obrigatorio'), 400
            return jsonify('Operação inválida'), 401
        return jsonify('Caixa Fechado'), 401

    def remover_despesas(self, bd, hd):
        id = bd.get('id')
        cr = hd.get('cr')
        perm = hd.get('perm')

        if id:
            if self.status_caixa(cr):
                if perm == 'ADMIN':
                    valor = self.cons("select valor from despesas where id = %s and cr = '%s'", (id, cr))
                    valor = valor[0] if valor else 0
                    self.query("update caixa_ab set valor = valor + %s where cr = '%s'", (valor, cr))
                    self.query("delete from despesas where id = %s and cr = '%s'", (id, cr))

                    return jsonify('Sucesso'), 200
                return jsonify('Operaçao invalida'), 401
            return jsonify('Caixa ainda fechado'), 401
        return jsonify('Id Obrigatorio'), 400

    # Pedidos =================================================== 
    def get_orders(self, cr):
        if cr:
            orders = []
            res = self.cons("select distinct p.cmd, f.nome, c.cliente, c.data from pedidos p inner join funcionarios f on f.matricula = p.funcionario inner join comandas c on c.cmd = p.cmd where p.cr = '%s' and p.status = 'SOLICITADO' ORDER BY c.data desc", cr, all=True)
            for cmd, func, cli, data in res:
                orders.append({
                    'cmd': cmd,
                    'func': func,
                    'cli': cli
                })
            return jsonify(orders), 200
        else: return jsonify('CR Obrigatorio')

    def new_order(self, bd, hd):
        cli = bd.get('cliente')
        cmd = bd.get('cmd')
        items = bd.get("items")
        cr = hd.get('cr')
        gc = hd.get('gc')
        mat = hd.get('mat')
        if cmd:
            if items:
                if cr and gc and mat:
                    if self.status_caixa(cr):
                        config = self.get_config(cr)
                        self.query('rollback')
                        res = self.cons("select cmd, valor_real from comandas where cmd = '%s' and cr = '%s'", (cmd, cr))
                        if res: cmd2, total = res
                        else: cmd2, total = [False, 0]
                        
                        for id in items:
                            quantidade = items[id]['quantidade']
                            res = self.cons("select p.valor, c.nome, p.nome, p.preparo from produtos p inner join categorias c on c.id = p.id_categoria where p.id = %s", id)
                            categoria = res[1] if res[1] else None
                            valor = res[0] if res[0] else None
                            nome = res[2] if res[2] else None
                            preparo = res[3] if res[3] else None
                            if valor and categoria and nome:
                                total += quantidade * valor
                                if categoria == 'COMBOS':
                                    items = self.cons("select c.id, c.quantidade, p.valor, p.nome from combo_items c inner join produtos p on p.id = c.id where c.cr = '%s' and c.combo_id = %s", (cr, id))
                                    valorCombo = valor
                                    for i in range(quantidade):
                                        id, quantidade, valor, nome = items
                                        if config['pedidos'] and preparo:
                                            self.query("insert into pedidos(id_produto, produto, quantidade, cmd, status, valor, funcionario, data, grupodecliente, cr) values(%s, '%s', %s, '%s', '%s', %s, %s, '%s', '%s', '%s')", (id, nome, quantidade, cmd, 'SOLICITADO', valorCombo, mat, now(self.get_fuso(cr)), gc, cr))
                                            self.query("update produtos set quantidade = quantidade - %s where id = %s", (quantidade, id))
                                        else:
                                            self.query("insert into pedidos(id_produto, produto, quantidade, cmd, status, valor, funcionario, data, grupodecliente, cr) values(%s, '%s', %s, '%s', '%s', %s, %s, '%s', '%s', '%s')", (id, nome, quantidade, cmd, 'ENTREGUE', valorCombo, mat, now(self.get_fuso(cr)), gc, cr))
                                            self.query("update produtos set quantidade = quantidade - %s where id = %s", (quantidade, id))
                                else:
                                    if config['pedidos'] and preparo:
                                        self.query("insert into pedidos(id_produto, produto, quantidade, cmd, status, valor, funcionario, data, grupodecliente, cr) values(%s, '%s', %s, '%s', '%s', %s, %s, '%s', '%s', '%s')", (id, nome, quantidade, cmd, 'SOLICITADO', (valor * quantidade), mat, now(self.get_fuso(cr)), gc, cr))
                                        self.query("update produtos set quantidade = quantidade - %s where id = %s", (quantidade, id))
                                    else:
                                        self.query("insert into pedidos(id_produto, produto, quantidade, cmd, status, valor, funcionario, data, grupodecliente, cr) values(%s, '%s', %s, '%s', '%s', %s, %s, '%s', '%s', '%s')", (id, nome, quantidade, cmd, 'ENTREGUE', (valor * quantidade), mat, now(self.get_fuso(cr)), gc, cr))
                                        self.query("update produtos set quantidade = quantidade - %s where id = %s", (quantidade, id))

                        if not cli: cli = 'NÃO INFORMADO'
                        if cmd2: self.query("update comandas set valor_real = %s where cmd = '%s' and cr = '%s'", (total, cmd, cr))
                        else: self.query("insert into comandas(cmd, valor_real, funcionario, data, cliente, grupodecliente, cr) values('%s', %s, %s, '%s', '%s', '%s', '%s')", (cmd, total, mat, now(self.get_fuso(cr)), cli.upper(), gc, cr))

                        return jsonify("Sucesso"), 200
                    return jsonify("Caixa fechado"), 401
                return jsonify("Confira as credenciais do Header!"), 401
            return jsonify("Adicione algum produto!"), 401
        return jsonify("Comanda ou Mesa vazia!"), 401

    def rm_order(self, bd, hd):
        cmd = bd.get('cmd')
        cr = hd.get('cr')
        if cmd:
            if self.status_caixa(cr):
                res = self.cons("select id_produto, quantidade, valor from pedidos where cmd = '%s' and cr = '%s' and status = 'SOLICITADO'", (cmd, cr), all=True)
                valor = 0
                for id, quant, vl in res:
                    valor += vl
                    self.query("update produtos set quantidade = quantidade + %s where id = %s", (quant, id))

                cmdAtiva = self.cons("select valor from pedidos where cmd = '%s' and cr = '%s' and status = 'ENTREGUE'", (cmd, cr))
                if not cmdAtiva: self.query("delete from comandas where cmd = '%s' and cr = '%s'", (cmd, cr))
                else:
                    res2 = self.query("update comandas set valor_real = valor_real - %s where cmd = '%s' and cr = '%s' returning valor_real", (valor, cmd, cr))
                self.query("update pedidos set status = 'CANCELADO' where cmd = '%s' and cr = '%s' and status = 'SOLICITADO'", (cmd, cr))
                return jsonify("Sucesso"), 200
            else: return jsonify("Caixa fechado!"), 401
        else: return jsonify("Comanda ou Mesa obrigatorio!"), 401

    def rm_order_with_id(self, bd, hd):
        cr = hd.get('cr')
        idPedido = bd.get('id')
        cmd = bd.get('cmd')

        if cr:
            if idPedido:
                if cmd:
                    # Confirmação do valor do pedido
                    res = self.cons("select valor, status from pedidos where id = %s and cr = '%s'", (idPedido, cr))
                    valor = res[0] if res else 0
                    status = res[1] if res else False

                    if status:
                        if status != 'CANCELADO':
                            # Devolutiva de produtos ao estoque
                            prods = self.cons("select id_produto, quantidade from pedidos where cr = '%s' and cmd = '%s'", (cr, cmd), all=True)
                            for id, quantidade in prods:
                                self.query("update produtos set quantidade = quantidade + %s where cr = '%s' and id = %s", (quantidade, cr, id))

                            # Atualiza o valor da comanda
                            res = self.query("update comandas set valor_real = valor_real - %s where cr = '%s' and cmd = '%s' returning valor_real", (valor, cr, cmd))
                            valor_real = res[0] if res else 0
                            print(valor_real, res)

                            if valor_real <= 0: 
                                print('Menor que ZERO')
                                self.query("delete from comandas where cr = '%s' and cmd = '%s'", (cr, cmd))
                                self.query("delete from pedidos where cr = '%s' and cmd = '%s'", (cr, cmd))

                            self.query("update pedidos set status = 'CANCELADO' where cr = '%s' and id = %s", (cr, idPedido))
                            return jsonify('Sucesso'), 200
                        return jsonify('Produto já cancelado'), 401
                    return jsonify('Status não encontrado'), 401
                return jsonify('Comanda Obrigatorio'), 401
            return jsonify('Id Obrigatorio'), 401
        return jsonify('CR Obrigatorio'), 401

    def rm_order_only(self, hd, bd):
        cr = hd.get('cr')
        cmd = bd.get('cmd')
        idp = bd.get('idp')
        id = bd.get('id')
        if cr:
            if id:
                if cmd:
                    res = self.cons("select valor, quantidade from pedidos where id = %s and cr = '%s'", (id, cr))
                    valor, quantidade = res
                    self.query("update comandas set valor_real = valor_real - %s where cmd = '%s' and cr = '%s'", (valor, cmd, cr))
                    self.query("update produtos set quantidade = quantidade + %s where id = %s and cr = '%s'", (quantidade, idp, cr))
                    self.query("update pedidos set status = 'CANCELADO' where id = %s", id)
                    return jsonify('Sucesso'), 200
                else: return jsonify('Comanda ou mesa obrigatoria'), 400
            else: return jsonify('Id obrigatorio'), 400
        else: return jsonify('Credenciais invalidas - Headers'), 400

    def set_leave_order(self, hd, bd):
        cr =  hd.get('cr')
        cmd = bd.get('cmd')

        if cmd:
            self.query("update pedidos set status = 'ENTREGUE' where cr = '%s' and cmd = '%s' and status = 'SOLICITADO'", (cr, cmd))
            return jsonify("Sucesso"), 200
        else: return jsonify("Id obrigatorio"), 400

    # Comandas ===================================================
    def get_cmds(self, cr):
        res = self.cons("select distinct c.cmd, c.cliente, c.valor_real, f.nome from comandas c inner join funcionarios f on c.funcionario = f.matricula where c.cr = '%s'", cr, all=True)
        cmds = []
        for cmd, cliente, valor, func in res:
            cmds.append({
                'cmd': cmd,
                'cliente': cliente,
                'valor': valor,
                'func': func
            })
        return jsonify(cmds), 200
    
    def rm_cmd(self, bd, hd):
        cmd = bd.get('cmd')
        cr = hd.get('cr')

        if cmd:
            if cr:
                if self.status_caixa(cr):
                    res = self.cons("select id_produto, quantidade from pedidos where cmd = '%s' and cr = '%s'", (cmd, cr), all=True)
                    for id, quant in res: self.query("update produtos set quantidade = quantidade + %s where id = %s", (quant, id))
                    self.query("delete from comandas where cr = '%s' and cmd = '%s'", (cr, cmd))
                    self.query("delete from pedidos where cmd = '%s' and cr = '%s' ", (cmd, cr))
                    return jsonify('Sucesso'), 200
                else: return jsonify('Caixa fechado!'), 401
            else: return jsonify('Credenciais incorretas - Headers'), 400     
        else: return jsonify('Comanda ou mesa obrigatorios!'), 400
   
    def get_cmd(self, hd, bd):
        cr = hd.get('cr')
        cmd = bd.get('cmd')

        if cr and cmd:
            cmd = unquote_plus(cmd)
            res = self.cons("select c.cmd, c.cliente, c.valor_real, f.nome, to_char(c.data, 'YYYY-MM-DD HH24:MI:SS') from comandas c inner join funcionarios f on f.matricula = c.funcionario where c.cr = '%s' and c.cmd = '%s'", (cr, cmd))
            if res: 
                cmd, cliente, valor_real, func, data = res

                prods = []
                cProd = self.cons("select p.id, p.produto, p.quantidade, p.valor, p.status, f.nome, to_char(p.data,'YYYY-MM-DD HH24:MI:SS') as dataa, pr.preparo from pedidos p inner join funcionarios f on f.matricula = p.funcionario inner join produtos pr on pr.id = p.id_produto where p.cmd = '%s' and p.cr = '%s' order by dataa desc", (cmd, cr), all=True)
                for idPedido, prod, quant, vl, st, funcP, data, preparo in cProd:
                    prods.append({
                        'id_pedido': idPedido,
                        'nome':prod,
                        'quant':quant,
                        'valor':vl,
                        'status':st,
                        'func':funcP,
                        'data': data,
                        'preparo': preparo
                    })

                r = {
                    'cmd':cmd,
                    'cli':cliente.capitalize(),
                    'total':valor_real,
                    'func':func.capitalize(),
                    'data':data,
                    'prods':prods
                }
                return jsonify(r), 200
            else: return jsonify('CMD Não encontrada'), 400
        else: return jsonify('Credenciais invalidas'), 400
    
    def close_cmd(self, bd, hd):
        valor = bd.get('valor', 0)

        debito = bd.get('debito', 0)
        debito = debito if debito else 0

        credito = bd.get('credito', 0)
        credito = credito if credito else 0

        pix = bd.get('pix', 0)
        pix = pix if pix else 0

        dinheiro = bd.get('dinheiro', 0)
        dinheiro = dinheiro if dinheiro else 0

        desconto = bd.get('desconto', 0)
        desconto = desconto if desconto else 0

        troco = bd.get('troco', 0)

        cliente = bd.get('cliente')
        cmd = bd.get('cmd')

        cr = hd.get('cr')
        gc = hd.get('gc')

        perm = hd.get('perm')

        if cr and cmd:
            if self.status_caixa(cr):
                if perm == 'ADMIN':
                    if valor:
                        # Confere se a comanda existe
                        res = self.cons("select valor_real, funcionario from comandas where cmd = '%s' and cr = '%s'", (cmd, cr))
                        if res:
                            valor_real = res[0]
                            funcionario = res[1]
                            if valor + desconto >= valor_real:
                                # Insere a venda no sistema
                                idVenda = self.query(
                                    """insert into vendas(
                                        cmd, 
                                        valor_real, 
                                        valor_pago, 
                                        cliente,
                                        grupodecliente, 
                                        cr, 
                                        data,
                                        status,
                                        debito, 
                                        credito,
                                        pix,
                                        dinheiro,
                                        desconto,
                                        troco,
                                        funcionario
                                    )
                                    values('%s', %s, %s, '%s', '%s', '%s', '%s', 'FINALIZADA', %s, %s, %s, %s, %s, %s, %s)
                                    returning id""", 
                                    (cmd, valor_real, (valor_real - desconto), cliente, gc,  cr, now(self.get_fuso(cr)), debito, credito, pix, dinheiro, desconto, troco, funcionario))

                                # Atualiza o valor do Caixa
                                if dinheiro > 0: self.query("update caixa_ab set valor = valor + %s where cr = '%s'", (dinheiro - desconto, cr))
                                
                                # Insere os produtos vendidos
                                prods = self.cons("select produto, quantidade, valor, funcionario, data from pedidos where cmd = '%s' and cr = '%s'", (cmd, cr), all=True)
                                for produto, quantidade, valor_produto, funcionario, data in prods:
                                    self.query("""
                                        insert into saidas 
                                            (id_venda,
                                            nome_produto,
                                            quantidade,
                                            valor,
                                            funcionario,
                                            data,
                                            cr)
                                        values(%s, '%s', %s, %s, %s, '%s', '%s')""",
                                        (idVenda[0], produto, quantidade, valor_produto, funcionario, data, cr)
                                    )
                                
                                # Limpa os pedidos e comandas
                                self.query("delete from comandas where cmd = '%s' and cr = '%s'", (cmd, cr))
                                self.query("delete from pedidos where cmd = '%s' and cr = '%s'", (cmd, cr))
                                return jsonify('Sucesso'), 200
                            return jsonify('Valor informado menor que o total!'), 400
                        return jsonify('Comanda não encontrada!'), 404
                    return jsonify('Valor Obrigatorio!'), 400
                return jsonify('Operacao Invalida!'), 401
            return jsonify('Caixa fechado!'), 401
        return jsonify('CR e Comanda obrigatorios!'), 400

    # Produtos ===================================================
    def consultar_produtos(self, hd):
        cr = hd.get('cr')
        prods = []
        res = self.cons("select p.id, p.sku, p.id_categoria, p.nome, c.nome, p.valor, p.quantidade, p.custo, to_char(p.data, 'YYYY-MM-DD HH24:MI:SS'), p.img, p.alerta, p.preparo from produtos p inner join categorias c on c.id = p.id_categoria where p.cr = '%s' order by c.nome", cr, all=True)

        for id, sku, id_ctg, nome, ctg, valor, quant, custo, data, img, alerta, preparo in res:
            if ctg != 'COMBOS':
                prods.append({
                    'id':id,
                    'sku': sku,
                    'nome':nome,
                    'categoria':ctg,
                    'valor': valor,
                    'quantidade': quant,
                    'alerta': alerta,
                    'custo': custo,
                    'data': data,
                    'img': img,
                    'id_ctg': id_ctg,
                    'preparo': preparo
                })

        return jsonify(prods),200

    def add_new_prod(self, bd, hd, files):
        cr = hd.get('cr')
        gc = hd.get('gc')
        
        nome = bd.get("nome")
        categoria = bd.get("categoria")
        custo = bd.get("custo")
        valor = bd.get("valor")
        quantidade = bd.get("quantidade")
        alerta = bd.get("alerta")
        preparo = bd.get('preparo')
        preparo = True if preparo == 'on' else False
        sku = bd.get('sku')

        if cr and gc:
            if nome:
                if categoria:
                    if custo:
                        if valor:
                            if quantidade:
                                idProd = self.query(
                                    """
                                        INSERT INTO produtos(
                                            nome, id_categoria, custo,
                                            valor, quantidade, alerta, 
                                            data, grupodecliente, cr, preparo
                                        )values(
                                            '%s', %s, %s, %s, %s, '%s',
                                            '%s', '%s', '%s', %s
                                        )returning id
                                    """, (
                                        nome.upper(), categoria, custo, valor,
                                        quantidade, alerta, 
                                        now(self.get_fuso(cr)), gc, cr, preparo

                                    )
                                )[0]

                                if files:
                                    img = files.get('img')
                                    if img:
                                        filename = f'prod_{idProd}.png'
                                        filepath = path.join(getcwd(), 'img/gourmet', filename)
                                        img.save(filepath)
                                    else: filename = 'blank.png'
                                else: filename = 'blank.png'

                                sku = sku if sku else idProd

                                self.query("update produtos set img = '%s', sku = '%s' where cr = '%s' and id = %s ", (filename, sku, cr, idProd))
                                
                                return jsonify('Sucesso'), 200
                            return jsonify('Quantidade obrigatoria'), 400
                        return jsonify('Valor obrigatorio'), 400
                    return jsonify('Custo obrigatorio'), 400
                return jsonify('Categoria obrigatoria'), 400
            return jsonify('Nome obrigatoria'), 400
        return jsonify('Credenciais invalidas'), 401

    def alter_prod(self, bd, hd, files):
        cr = hd.get('cr')
        gc = hd.get('gc')
        
        id = bd.get('idprod')
        nome = bd.get("nome")
        categoria = bd.get("categoria")
        custo = bd.get("custo")
        valor = bd.get("valor")
        quantidade = bd.get("quantidade")
        alerta = bd.get("alerta")
        sku = bd.get('sku')
        preparo = bd.get('preparo')
        preparo = True if preparo == 'on' else False

        if cr:
            if id:
                self.query("""
                           update produtos set
                           nome = '%s', id_categoria = %s,
                           custo = %s, valor = %s,
                           quantidade = %s, alerta = %s, sku = '%s', 
                           preparo = %s
                           where cr = '%s' and id = %s 
                           """,(
                               nome, categoria, custo, 
                               valor, quantidade, alerta, 
                               sku, preparo, cr, id
                           ))
                if files:
                    img = files.get('img')
                    if img:
                        filename = f'prod_{id}.png'
                        filepath = path.join(getcwd(), 'img/gourmet', filename)
                        img.save(filepath)
                        self.query("update produtos set img = '%s' where cr = '%s' and id = %s", (filename, cr, id))

                return jsonify('Sucesso'), 200
            return jsonify('Id Obrigatorio'), 400
        return jsonify('Credenciais invalidas'), 401

    def rm_prod(self, bd, hd):
        id = bd.get('id')
        cr = hd.get('cr')

        if cr:
            if id:
                self.query("delete from produtos where id = %s and cr = '%s'", (id, cr))
                return jsonify('Sucesso'), 200
            return jsonify('Id Obrigatorio'), 400
        return jsonify('Credenciais Invalidas'), 401

    def put_prod(self, bd, hf):
        ...

    # Produtos dos Pedidos ===================================================
    def add_prod(self, bd): # ADICIONAR PROD AOS PEDIDOS - EVITA VENDA DE ESTOQUE NEGATIVO
        id = bd.get("id")
        if id:
            self.query("update produtos set quantidade = quantidade + 1 where id = %s", id, commit=False)
            return jsonify("Sucesso"), 200
        else: return jsonify("Id obrigatorio"), 400
    
    def rmv_prod(self, bd): # REMOVER PROD AOS PEDIDOS - EVITA VENDA DE ESTOQUE NEGATIVO
        id = bd.get("id")
        if id:
            res = self.cons("select quantidade from produtos where id = %s and quantidade > 0", id)
            if res:
                self.query("update produtos set quantidade = quantidade - 1 where id = %s returning quantidade", id, commit=False)
                return jsonify(res), 200
            else: return jsonify("Quantidade Zerada"), 401
        else: return jsonify("Id obrigatorio"), 400

    def prods_pedidos(self, bd, cr):
        cmd = bd.get("cmd")
        prods = []
        if cmd:
            res = self.cons("select id, id_produto, produto, quantidade, status from pedidos where cr= '%s' and cmd = '%s' and status = 'SOLICITADO'", (cr, cmd), all=True)
            for id, idp, prod, quant, status in res:
                prods.append({
                    'id': id,
                    'idp': idp,
                    'prod': prod,
                    'quant': quant,
                    'status': status
                })
            return jsonify(prods), 200
        else: return jsonify('CR Obrigatorio!'), 400

    # Vendas ===================================================
    def get_vendas(self, hd):
        cr = hd.get('cr')
        if cr:
            res = self.cons("select v.id, v.cmd, v.valor_real, v.valor_pago, f.nome, v.cliente, to_char(v.data, 'YYYY-MM-DD HH24:MI:SS'), v.status from vendas v inner join funcionarios f on f.matricula = v.funcionario where v.cr = '%s' order by v.data desc", (cr, ), all=True)
            dd = []
            for id, cmd, valor_real, valor_pago, func, cli, data, status in res:
                dd.append({
                    'id': id,
                    'cmd': cmd,
                    'valor_real': valor_real,
                    'valor_pago': valor_pago,
                    'funcionario': func,
                    'cliente': cli,
                    'data': data,
                    'status': status
                })
            return jsonify(dd), 200
        else: return jsonify('CR Obrigatorio!'), 400

    def rm_venda(self, bd, hd):
        cr = hd.get('cr')
        perm = hd.get('perm')

        id = bd.get('id')

        if cr:
            if perm == 'ADMIN':
                if id:
                    dinheiro = self.cons("select dinheiro from vendas where id = %s and cr = '%s'", (id, cr))
                    dinheiro = dinheiro[0] if dinheiro else False
                    if dinheiro: self.query("update caixa_ab set valor = valor - %s where cr = '%s'", (dinheiro, cr))

                    self.query("delete from vendas where id = %s and cr = '%s'", (id, cr))
                    self.query("delete from saidas where id_venda = %s and cr = '%s'", (id, cr))

                    return jsonify('Sucesso'), 200
                return jsonify('ID Obrigatorio'), 400
            return jsonify('Operação Inválida'), 401
        return jsonify('CR Obrigatório'), 400

    def set_venda(self, bd, hd):
        cr = hd.get('cr')
        gc = hd.get('gc')
        mat = hd.get('mat')

        valor = bd.get('valor')
        cart = bd.get("cart")
        desconto = bd.get("desconto")
        desconto = desconto if desconto else 0
        mt_pg = bd.get("mt_pg")

        debito = 0
        credito = 0
        pix = 0
        dinheiro = 0

        if cr and gc:
            if self.status_caixa(cr):
                if valor:
                    if cart:
                        if mt_pg:
                            match mt_pg:
                                case 'debito': debito = valor
                                case 'credito': credito = valor
                                case 'pix': pix = valor
                                case 'dinheiro': dinheiro = valor

                            idVenda = self.query(
                                """
                                    insert into vendas(
                                        cmd, valor_real,
                                        valor_pago, cliente,
                                        data, funcionario,
                                        cr, grupodecliente, status, debito, 
                                        credito, pix, dinheiro, 
                                        desconto, troco
                                    )
                                    values(
                                        'CAIXA', %s, %s,
                                        'Não Informado', '%s',
                                        %s, '%s', '%s', 'FINALIZADA',
                                        %s, %s, %s, %s, %s, 0 
                                    )
                                    returning id
                                """,(
                                    valor, (valor - desconto), now(self.get_fuso(cr)),
                                    mat, cr, gc, debito, credito,
                                    pix, dinheiro, desconto
                                )
                            )[0]


                            for id in cart:
                                quant = cart[id]['quantidade']
                                nome = cart[id]['nome']
                                valorProd = cart[id]['valor']
                                categ = self.cons("select c.nome from produtos p inner join categorias c on c.id = p.id_categoria where p.id = %s and p.cr = '%s'", (id, cr))
                                categ = categ[0] if categ else None
                                if categ == 'COMBOS':
                                    items = self.cons("select c.id, c.quantidade, p.valor, p.nome from combo_items c inner join produtos p on p.id = c.id where c.cr = '%s' and c.combo_id = %s", (cr, id))
                                    for i in range(quant):
                                        idProd, quantidade, valorP, nomeP = items
                                        self.query("update produtos set quantidade = quantidade - %s where id = %s and cr = '%s'", (quantidade, idProd, cr))
                                        for _ in range(quantidade):
                                            self.query(
                                                """
                                                    insert into saidas(
                                                        id_venda, nome_produto, quantidade,
                                                        valor, funcionario, data, cr
                                                    )
                                                    values(
                                                        %s, '%s', %s, %s, %s, '%s', '%s'
                                                    )
                                                """,(
                                                    idVenda, nomeP, 1, 
                                                    valorP, mat, 
                                                    now(self.get_fuso(cr)), cr
                                                )
                                            )
                                else:
                                    self.query("update produtos set quantidade = quantidade - %s where id = %s and cr = '%s'", (quant, id, cr))
                                    for _ in range(quant):
                                        self.query(
                                            """
                                                insert into saidas(
                                                    id_venda, nome_produto, quantidade,
                                                    valor, funcionario, data, cr
                                                )
                                                values(
                                                    %s, '%s', %s, %s, %s, '%s', '%s'
                                                )
                                            """,(
                                                idVenda, nome, quant, 
                                                valorProd, mat, 
                                                now(self.get_fuso(cr)), cr
                                            )
                                        )

                            if dinheiro > 0: self.query("update caixa_ab set valor = valor + %s where cr = '%s'", ((valor - desconto), cr))
                            return jsonify("Sucesso"), 200
                        return jsonify("Metodo de Pagamento invalido"), 400
                    return jsonify("Adicione um item em seu carrinho!"), 400
                return jsonify("Valor Obrigatório"), 400
            return jsonify("Caixa Fechado"), 401
        return jsonify("Credenciais invalidas"), 400

    # Saidas ===================================================
    def get_saida(self, bd, hd):
        cr = hd.get('cr')
        id = bd.get('id')
        saida = []
        if cr:
            if id:
                res = self.cons("select nome_produto, quantidade, valor from saidas where id_venda = %s and cr = '%s'", (id, cr), all=True)
                for nome, quant, valor in res:
                    saida.append({
                        'nome': nome,
                        'quantidade': quant,
                        'valor': valor
                    })
                return jsonify(saida), 200
            return jsonify('Id Obrigatorio'), 400
        return jsonify('CR Obrigatorio'), 400

    # Categorias ===================================================
    def get_categories(self, hd):
        cr = hd.get('cr')
        gc = hd.get('gc')
        ctgs = []
        if cr:
            config = self.get_config(cr)
            if config['combos']:
                combo = self.cons("select nome from categorias where cr = '%s' and nome = 'COMBOS'", cr)
                print(combo)
                if not combo: self.query("insert into categorias (nome, cr, grupodecliente) values('COMBOS', '%s', '%s')", (cr, gc))
            res = self.cons("select id, nome from categorias where cr = '%s'", cr, all=True)
            for id, nome in res:
                if nome != 'COMBOS':
                    ctgs.append({
                        'id': id,
                        'nome': nome
                    })
            return jsonify(ctgs), 200
        return jsonify("Credenciais invalidas"), 401

    def add_categories(self, bd, hd):
        cr = hd.get('cr')
        gc = hd.get('gc')

        nome = bd.get('nome')

        if cr:
            if nome:    
                if nome != 'COMBOS':
                    self.query("insert into categorias(nome, cr, grupodecliente) values('%s','%s','%s')", (nome, cr, gc))
                    return jsonify("Sucesso"), 200
                return jsonify("Combos não permitido, adicione pela aba 'COMBOS'"), 401
            return jsonify("Nome obrigatorio"), 400
        return jsonify("Credenciais invalidas"), 401

    def update_categories(self, bd, hd):
        cr = hd.get('cr')
        nome = bd.get('nome')

        if cr:
            if nome:    
                self.query("update categorias set nome = '%s' where cr = '%s'", (nome, cr))
                return jsonify("Sucesso"), 200
            return jsonify("Nome obrigatorio"), 400
        return jsonify("Credenciais invalidas"), 401

    def rm_categories(self, bd, hd):
        cr = hd.get('cr')

        id = bd.get('id')

        if cr:
            if id:    
                self.query("delete from categorias where cr = '%s' and id = %s", (cr, id))
                return jsonify("Sucesso"), 200
            return jsonify("Id obrigatorio"), 400
        return jsonify("Credenciais invalidas"), 401

    # Combos ===================================================
    def get_combos(self, bd, hd):
        cr = hd.get('cr')
        gc = hd.get('gc')
        combos = []
        if cr:
            res = self.cons(
                """
                    select c.id, p.nome, p.valor, p.img, c.ativo, ct.nome
                    from combos c
                    inner join produtos p
                    on c.id = p.id
                    inner join categorias ct
                    on ct.id = p.id_categoria
                    where c.cr = '%s' 
                    and c.grupodecliente = '%s'
                """, 
                (cr, gc), 
                all=True
            )
            for id, nomeCombo, valor, img, ativo, cate in res:
                prods = self.cons("select c.id, p.nome, c.quantidade from combo_items c inner join produtos p on p.id = c.id where c.cr = '%s' and c.grupodecliente = '%s' and c.combo_id = %s", (cr, gc, id), all=True)
                list_of_prods = [] # Lista de Produtos do COMBO
                for prod in prods: # Adicona os produtos do COMBO a lista
                    idProd, nome, quantidade = prod
                    list_of_prods.append({
                        'id': idProd,
                        'nome': nome,
                        'quantidade': quantidade
                    })
                combos.append({
                    'id': id,
                    'nome': nomeCombo,
                    'valor': valor,
                    'img': img,
                    'ativo': ativo,
                    'items': list_of_prods
                })
            return jsonify(combos)
        return jsonify("Credenciais Invalidas"), 401

    def set_combo(self, bd, hd, files):
        cr = hd.get('cr')
        gc = hd.get('gc')
        perm = hd.get('perm')

        nome = bd.get('nome')
        sku = bd.get('sku')
        custo = bd.get('custo')
        valor = bd.get('valor')
        items = bd.get('items')
        preparo = bd.get('preparo')
        preparo = True if preparo == 'on' else False

        if cr and gc and perm == 'ADMIN':
            id_ctg = self.cons("select id from categorias where nome = 'COMBOS' and cr = '%s'", (cr))[0]
            if not id_ctg: id_ctg = self.query("insert into categorias(nome, cr, grupodecliente) values('COMBOS', '%s', '%s') returning id", (cr, gc))[0]
                
            # Cria o produto do combo
            idProd = self.query("""
                insert into produtos(
                    nome, id_categoria, 
                    custo, valor, quantidade, 
                    alerta, data, grupodecliente, cr, preparo
                ) 
                values(
                    '%s', %s, %s, %s, 0, 0, 
                    '%s', '%s', '%s', %s
                ) returning id
            """, (
                nome.upper(), id_ctg, custo, 
                valor, now(self.get_fuso(cr)), 
                gc, cr, preparo
            ))[0]
            sku = sku if sku else idProd

            if files:
                img = files.get('img')
                if img:
                    filename = f'prod_{idProd}.png'
                    filepath = path.join(getcwd(), 'img/gourmet', filename)
                    img.save(filepath)
                else: filename = 'blank.png'
            else: filename = 'blank.png'
            sku = sku if sku else idProd

            # Atualiza o SKU/IMG do produto
            self.query("update produtos set img = '%s', sku = '%s' where cr = '%s' and id = %s ", (filename, sku, cr, idProd))
            
            # Cria o combo 
            self.query("insert into combos(id, ativo, grupodecliente, cr) values(%s, true, '%s', '%s')", (idProd, gc, cr))

            # Adiciona os itens ao combo
            items = loads(items)
            for id in items:
                quantidade = items[id]['quantidade']
                self.query("insert into combo_items(id, quantidade, combo_id, grupodecliente, cr) values(%s, %s, %s, '%s', '%s')", (id, quantidade, idProd, gc, cr))
            

            return jsonify("Sucesso"), 200
        return jsonify("Credenciais Invalidas"), 401

    def update_combo(self, bd, hd):
        ...

    def remove_combo(self, bd, hd):
        id = bd.get('id')

        cr = hd.get('cr')
        gc = hd.get('gc')
        if cr and gc:
            if id:
                self.query("delete from combo_items where combo_id = %s and cr = '%s' and grupodecliente = '%s'", (id, cr, gc))
                self.query("delete from combos where id = %s and cr = '%s' and grupodecliente = '%s'", (id, cr, gc))
                self.query("delete from produtos where id = %s and cr = '%s'", (id, cr))
                return jsonify("Sucesso"), 200
            return jsonify("Id obrigatorio"), 400
        return jsonify("Credenciais invalidas"), 401

    # Configurações ===================================================
    def get_config(self, cr): # Função para uso INTERNO apenas!
        res = self.cons("select imprimir, pedidos, comandas, estoque, combos, fuso, email from config where cr = '%s'", (cr))
        imp, ped, cmd, est, cmb, fuso, email = res
        config = {
            'imprimir': imp,
            'pedidos': ped,
            'comandas': cmd,
            'estoque': est,
            'combos': cmb,
            'fuso': fuso,
            'email': email
        }
        return config

    def get_config_ext(self, hd): # Função para uso EXTERNO apenas!
        cr = hd.get('cr')

        if cr:
            res = self.cons("select imprimir, pedidos, comandas, estoque, combos, fuso, email from config where cr = '%s'", (cr))
            imp, ped, cmd, est, cmb, fuso, email = res
            config = {
                'imprimir': imp,
                'pedidos': ped,
                'comandas': cmd,
                'estoque': est,
                'combos': cmb,
                'fuso': fuso,
                'email': email
            }
            return jsonify(config), 200
        return jsonify('Credenciais invalidas'), 401

    def update_config(self, bd, hd):
        cr = hd.get('cr')
        config = bd.get('config')
        config_value = bd.get('value')

        if cr:
            match config:
                case 'fuso': self.query("update config set fuso = %s where cr = '%s'", (config_value, cr))
                case 'imp': self.query("update config set imprimir = %s where cr = '%s'", (config_value, cr))
                case 'pedidos': self.query("update config set pedidos = %s where cr = '%s'", (config_value, cr))
                case 'comandas': self.query("update config set comandas = %s where cr = '%s'", (config_value, cr))
                case 'estoque': self.query("update config set estoque = %s where cr = '%s'", (config_value, cr))
                case 'combos': self.query("update config set combos = %s where cr = '%s'", (config_value, cr))
                case 'logo': self.query("update config set logo = %s where cr = '%s'", (config_value, cr))
                case 'email': self.query("update config set email = '%s' where cr = '%s'", (config_value, cr))
            return jsonify("Sucesso"), 200
        return jsonify("Credenciais invalidas"), 401

    # Funcionarios ===================================================
    def get_employees(self, hd):
        cr = hd.get('cr')
        gc = hd.get('gc')
        employees = []
        if cr and gc:
            res = self.cons("select matricula, nome, permissao from funcionarios where cr = '%s' and grupodecliente = '%s'", (cr, gc), all=True)
            for mat, nome, perm in res:
                employees.append({
                    'mat': mat,
                    'nome': nome,
                    'perm': perm,
                })
            return jsonify(employees), 200
        return jsonify("Credemciais invalidas"), 401

    def create_employee(self, bd, hd):
        cr = hd.get('cr')
        gc = hd.get('gc')

        nome = bd.get('nome')
        pwd = bd.get('pwd')
        perm = bd.get('perm')

        if cr and gc:
            if nome:
                if pwd:
                    if perm:
                        hashPwd = sha256(pwd.encode()).hexdigest()
                        self.query("""
                            INSERT INTO funcionarios(
                                nome, hash, permissao, cr, grupodecliente
                            )
                            VALUES (
                                '%s', '%s', '%s', '%s', '%s'
                            )
                        """, (nome, hashPwd, perm, cr, gc))
                        return jsonify("Sucesso!"), 200
                    return jsonify("Permissao Obrigatoria"), 400
                return jsonify("Senha Obrigatoria"), 400
            return jsonify("Nome Obrigatorio"), 400
        return jsonify("Credenciais Invalidas"), 401
        
    def update_perm_employee(self, bd, hd):
        cr = hd.get('cr')
        gc = hd.get('gc')
        mat = bd.get("mat")

        if cr and gc:
            if mat:
                self.query("update funcionarios set permissao = 'ADMIN' where cr = '%s' and grupodecliente = '%s' and matricula = '%s'", (cr, gc, mat))
                return jsonify("Sucesso"), 200
            return jsonify("Matricula Invalida"), 401
        return jsonify("Credenciais Invalidas"), 401

    def delete_employee(self, bd, hd):
        cr = hd.get('cr')
        gc = hd.get('gc')

        mat = bd.get("mat")

        if cr and gc:
            if mat:
                self.query("delete from funcionarios where matricula = %s and cr = '%s' and grupodecliente = '%s'", (mat, cr, gc))
                return jsonify("Sucesso"), 200
            return jsonify("Matricula obrigatoria"), 400
        return jsonify("Credenciais Invalidas"), 401

    def get_name_employe(self, mat):
        res = self.cons("select nome from funcionarios where matricula = %s", mat)[0]
        return res

    # Relatorios ===================================================
    def get_info(self, bd, hd):
        cr = hd.get('cr')
        filter = bd.get('filter')
        rlt = {
            'st_vendas': {},
            'pagamentos': 
                {
                    'debito': 0,
                    'credito': 0,
                    'pix': 0,
                    'dinheiro': 0
                },
            'ticket_medio': 0,
            'faturamento': 0,
            'ticket_pc': 0,
            'estoque_alerta': [],
            'produtos_zerados': [],
        }
        if cr and filter:
            match filter:
                case 'dia':
                    ontem = (now(self.get_fuso(cr)) - timedelta(days=1)).strftime('%Y-%m-%d')
                    hoje = now(self.get_fuso(cr)).strftime('%Y-%m-%d')
                    hojeDia = int(now(self.get_fuso(cr)).strftime('%d'))

                    res = self.cons("select distinct to_char(data, 'DD') as dia, sum(valor_pago) as total from vendas where cr = '%s' and to_char(data, 'YYYY-MM-DD') >= '%s' group by dia order by dia desc", (cr, ontem), all=True)
                    dod = {
                        "hoje": 0,
                        "ontem": 0
                    }
                    for dia, total in res: 
                        if int(dia) == hojeDia: dod["hoje"] += total
                        else: dod["ontem"] += total
                    rlt['st_vendas'] = dod

                    pags = self.cons("select debito, credito, pix, dinheiro, desconto from vendas where cr = '%s' and to_char(data, 'YYYY-MM-DD') >= '%s'", (cr, hoje), all=True)
                    for debito, credito, pix, dinheiro, desconto in pags:
                        # Confere se tem desconto
                        if desconto > 0:
                            pag = {
                                'debito': debito,
                                'credito': credito,
                                'pix': pix,
                                'dinheiro': dinheiro
                            }

                            # Pega o maior valor estipulando
                            maior = max(pag, key=pag.get)
                            pag[maior] -= desconto

                            # Soma os valores
                            debito = pag['debito']
                            credito = pag['credito']
                            pix = pag['pix']
                            dinheiro = pag['dinheiro']

                        rlt['pagamentos']['debito'] += debito
                        rlt['pagamentos']['credito'] += credito
                        rlt['pagamentos']['pix'] += pix
                        rlt['pagamentos']['dinheiro'] += dinheiro

                    tc = self.cons("select sum(valor_pago), count(valor_pago) from vendas where cr = '%s' and to_char(data, 'YYYY-MM-DD') >= '%s'", (cr, hoje))
                    faturamento = tc[0] if tc[0] else 0
                    numV = tc[1] if tc[1] else 0
                    tcmedio = faturamento / numV if numV > 0 else 0
                    tcpc = self.cons("select distinct id_venda, count(distinct nome_produto) from saidas where cr = '%s' and to_char(data, 'YYYY-MM') >= '%s' group by id_venda", (cr, hoje), all=True)
                    total_itens = sum(qtd for _, qtd in tcpc)
                    num_cupons = len(tcpc)
                    tcpc = total_itens / num_cupons if num_cupons > 0 else 0

                    rlt['ticket_pc'] = round(tcpc, 2)
                    rlt['ticket_medio'] = round(tcmedio, 2)
                    rlt['faturamento'] = faturamento
                
                case 'mes':
                    passado = (now(self.get_fuso(cr)) - timedelta(30)).strftime('%Y-%m')
                    atual = now(self.get_fuso(cr)).strftime('%Y-%m')
                    mesA = int(now(self.get_fuso(cr)).strftime('%m'))

                    res = self.cons("select distinct to_char(data, 'MM') as mes, sum(valor_pago) as total from vendas where cr = '%s' and to_char(data, 'YYYY-MM') >= '%s' group by mes order by mes desc", (cr, passado), all=True)
                    dod = {
                        meses[mesA]: 0,
                        meses[mesA-1 if mesA > 1 else 12]: 0
                    }
                    for mes, total in res: 
                        mes = int(mes)
                        if mes == mesA: dod[meses[mes]] += total
                        else: dod[meses[mes]] += total
                    rlt['st_vendas'] = dod

                    pags = self.cons("select debito, credito, pix, dinheiro, desconto from vendas where cr = '%s' and to_char(data, 'YYYY-MM') >= '%s'", (cr, atual), all=True)
                    for debito, credito, pix, dinheiro, desconto in pags:
                        # Confere se tem desconto
                        if desconto > 0:
                            pag = {
                                'debito': debito,
                                'credito': credito,
                                'pix': pix,
                                'dinheiro': dinheiro
                            }

                            # Pega o maior valor estipulando
                            maior = max(pag, key=pag.get)
                            pag[maior] -= desconto

                            # Soma os valores
                            debito = pag['debito']
                            credito = pag['credito']
                            pix = pag['pix']
                            dinheiro = pag['dinheiro']

                        rlt['pagamentos']['debito'] += debito
                        rlt['pagamentos']['credito'] += credito
                        rlt['pagamentos']['pix'] += pix
                        rlt['pagamentos']['dinheiro'] += dinheiro

                    tc = self.cons("select sum(valor_pago), count(valor_pago) from vendas where cr = '%s' and to_char(data, 'YYYY-MM') >= '%s'", (cr, atual))
                    faturamento = tc[0] if tc[0] else 0
                    numV = tc[1] if tc[1] else 0
                    tcmedio = faturamento / numV if numV > 0 else 0
                    tcpc = self.cons("select distinct id_venda, count(distinct nome_produto) from saidas where cr = '%s' and to_char(data, 'YYYY-MM') >= '%s' group by id_venda", (cr, atual), all=True)
                    total_itens = sum(qtd for _, qtd in tcpc)
                    num_cupons = len(tcpc)
                    tcpc = total_itens / num_cupons if num_cupons > 0 else 0
                    rlt['ticket_pc'] = round(tcpc, 2)
                    rlt['ticket_medio'] = round(tcmedio, 2)
                    rlt['faturamento'] = faturamento
                
                case 'ano': 
                    passado = (now(self.get_fuso(cr)) - timedelta(365)).strftime('%Y')
                    atual = now(self.get_fuso(cr)).strftime('%Y')
                    anoa = int(atual)

                    res = self.cons("select distinct to_char(data, 'YYYY') as ano, sum(valor_pago) as total from vendas where cr = '%s' and to_char(data, 'YYYY') >= '%s' group by ano order by ano desc", (cr, passado), all=True)
                    dod = {
                        passado: 0,
                        atual: 0
                    }
                    for ano, total in res: 
                        ano = int(ano)
                        if ano == anoa: dod[atual] += total
                        else: dod[passado] += total
                    rlt['st_vendas'] = dod

                    pags = self.cons("select debito, credito, pix, dinheiro, desconto from vendas where cr = '%s' and to_char(data, 'YYYY') >= '%s'", (cr, atual), all=True)

                    for debito, credito, pix, dinheiro, desconto in pags:
                        # Confere se tem desconto
                        if desconto > 0:
                            pag = {
                                'debito': debito,
                                'credito': credito,
                                'pix': pix,
                                'dinheiro': dinheiro
                            }

                            # Pega o maior valor estipulando
                            maior = max(pag, key=pag.get)
                            pag[maior] -= desconto

                            # Soma os valores
                            debito = pag['debito']
                            credito = pag['credito']
                            pix = pag['pix']
                            dinheiro = pag['dinheiro']

                        rlt['pagamentos']['debito'] += debito
                        rlt['pagamentos']['credito'] += credito
                        rlt['pagamentos']['pix'] += pix
                        rlt['pagamentos']['dinheiro'] += dinheiro

                    tc = self.cons("select sum(valor_pago), count(valor_pago) from vendas where cr = '%s' and to_char(data, 'YYYY') >= '%s'", (cr, atual))
                    faturamento = tc[0] if tc[0] else 0
                    numV = tc[1] if tc[1] else 0
                    tcmedio = faturamento / numV if numV > 0 else 0
                    tcpc = self.cons("select distinct id_venda, count(distinct nome_produto) from saidas where cr = '%s' and to_char(data, 'YYYY') >= '%s' group by id_venda", (cr, atual), all=True)
                    total_itens = sum(qtd for _, qtd in tcpc)
                    num_cupons = len(tcpc)
                    tcpc = total_itens / num_cupons if num_cupons > 0 else 0
                    rlt['ticket_pc'] = round(tcpc, 2)
                    rlt['ticket_medio'] = round(tcmedio, 2)
                    rlt['faturamento'] = faturamento
            limitado = self.cons("select p.nome, p.quantidade from produtos p inner join categorias c on c.id = p.id_categoria where p.cr = '%s' and p.quantidade <= p.alerta and c.nome <> 'COMBOS' order by p.quantidade asc", (cr), all=True)
            for nome, quant in limitado:
                if quant > 0:
                    rlt['estoque_alerta'].append({
                        'nome': nome,
                        'quantidade': quant
                    })
                else:
                    rlt['produtos_zerados'].append({
                        'nome': nome,
                        'quantidade': quant
                    })
            return jsonify(rlt), 200
        return jsonify('Credenciais invalidas'), 401
    
    def create_rl_vendas(self, data:str, flt:str, cr:str, arquivos:list, opr='='):
        if flt:
            with self.engine.connect() as conn:
                # Todas as vendas do mes
                cons = """select
                        cmd,
                        cliente,
                        to_char(data, 'DD/MM/YYYY HH24:MI') as "Data de Venda",
                        to_char(valor_pago, 'R$ FM9G999G999D00') as "Valor Total",
                        to_char(desconto, 'R$ FM9G999G9990D00') as "Desconto",
                        to_char(sum(valor_pago) - sum(desconto), 'R$ FM9G999G999D00') as "Valor Pago",
                        func as "Atendente"
                    from vw_vendas 
                    where data %s '%s'
                    and cr = '%s'
                    group by cmd, cliente, data, valor_pago, desconto, func
                    order by data desc;
                """ % (opr, data, cr)
                df_vendas = read_sql(cons, conn)
                df_vendas.to_excel(arquivos[0], index=False)

                # Info das Vendas
                cons = """select
                    count(distinct to_char(data, 'DD-MM')) as "Dias Trabalhados",
                    count(distinct id) as "Quantidade Vendas",
                    to_char(AVG(valor_pago - desconto), 'R$ FM9G999G999D00') as "Ticket Médio",
                    case 
                        when sum(valor_pago) > 0 then to_char(sum(valor_pago), 'R$ FM9G999G999D00')
                        else to_char(0, 'R$ FM9G999G9990D00') end as "Total Bruto",
                    case 
                        when sum(valor_pago) - sum(desconto) > 0 then to_char(sum(valor_pago) - sum(desconto), 'R$ FM9G999G999D00')
                        else to_char(0, 'R$ FM9G999G9990D00') end as "Total Liquido"
                    from vw_vendas
                    where data %s '%s'
                    and cr = '%s';"""%(opr, data, cr)
                df_infos = read_sql(cons, conn)
                df_infos.to_excel(arquivos[1], index=False)

                # Vendas por Funcionarios
                cons = """select 
                    func as funcionario,
                    count(func) as vendas
                from vw_vendas
                where data %s '%s'
                and cr = '%s'
                group by func;"""%(opr, data, cr)
                df_func = read_sql(cons, conn)
                df_func.to_excel(arquivos[2], index=False)

                # Saidas de Produtos
                cons = """select
                    nome_produto,
                    sum(quantidade) as "Total",
                    to_char(sum(valor), 'R$ FM9G999G999D00') as "Valor Total"
                from saidas
                where data %s '%s'
                and cr = '%s'
                group by nome_produto
                order by "Total" desc;""" % (opr, data, cr)
                df_saidas = read_sql(cons, conn)
                df_saidas.to_excel(arquivos[3], index=False)

                wb_final = Workbook()
                wb_final.remove(wb_final.active)
                return wb_final
        else:
            with self.engine.connect() as conn:
                # Todas as vendas do mes
                cons = """select
                        cmd,
                        cliente,
                        to_char(data, 'DD/MM/YYYY HH24:MI') as "Data de Venda",
                        to_char(valor_pago, 'R$ FM9G999G999D00') as "Valor Total",
                        to_char(desconto, 'R$ FM9G999G9990D00') as "Desconto",
                        to_char(sum(valor_pago) - sum(desconto), 'R$ FM9G999G999D00') as "Valor Pago",
                        func as "Atendente"
                    from vw_vendas 
                    where to_char(data, '%s') %s '%s'
                    and cr = '%s'
                    group by cmd, cliente, data, valor_pago, desconto, func
                    order by data desc;
                """ % (flt, opr, data, cr)
                df_vendas = read_sql(cons, conn)
                df_vendas.to_excel(arquivos[0], index=False)

                # Info das Vendas
                cons = """select
                    count(distinct to_char(data, 'DD-MM')) as "Dias Trabalhados",
                    count(distinct id) as "Quantidade Vendas",
                    to_char(AVG(valor_pago - desconto), 'R$ FM9G999G999D00') as "Ticket Médio",
                    case 
                        when sum(valor_pago) > 0 then to_char(sum(valor_pago), 'R$ FM9G999G999D00')
                        else to_char(0, 'R$ FM9G999G9990D00') end as "Total Bruto",
                    case 
                        when sum(valor_pago) - sum(desconto) > 0 then to_char(sum(valor_pago) - sum(desconto), 'R$ FM9G999G999D00')
                        else to_char(0, 'R$ FM9G999G9990D00') end as "Total Liquido"
                    from vw_vendas
                    where to_char(data, '%s') %s '%s'
                    and cr = '%s';"""%(flt, opr, data, cr)
                df_infos = read_sql(cons, conn)
                df_infos.to_excel(arquivos[1], index=False)

                # Vendas por Funcionarios
                cons = """select 
                    func as "Funcionario",
                    count(func) as "Total de Vendas"
                from vw_vendas
                where to_char(data, '%s') %s '%s'
                and cr = '%s'
                group by func;"""%(flt, opr, data, cr)
                df_func = read_sql(cons, conn)
                df_func.to_excel(arquivos[2], index=False)

                # Saidas de Produtos
                cons = """select
                    nome_produto,
                    count(nome_produto) as "Total",
                    to_char(sum(valor), 'R$ FM9G999G999D00') as "Valor Total"
                from saidas
                where to_char(data, '%s') %s '%s'
                and cr = '%s'
                group by nome_produto
                order by "Total" desc;""" % (flt, opr, data, cr)
                df_saidas = read_sql(cons, conn)
                df_saidas.to_excel(arquivos[3], index=False)

                wb_final = Workbook()
                wb_final.remove(wb_final.active)
                return wb_final
            
    def get_rl_vendas(self, bd, hd):
        cr = hd.get('cr')
        valorF = bd.get('valor', None)
        filter = bd.get('filter', None)
        opr = bd.get('operador', '=')
        flt = bd.get('flt', False)
        arquivos = [
            'src/vendas.xlsx',
            'src/infos.xlsx',
            'src/func.xlsx',
            'src/saidas.xlsx'
        ]
        
        if cr:
            if valorF and filter:
                match filter:
                    case 'dia':
                        if flt:
                            data = now().strftime('%d_%m_%Y')
                            wb_final = self.create_rl_vendas(valorF, flt, cr, arquivos, opr)
                        else:
                            dia = int(valorF)
                            if dia < 10: dia = f"0{dia}"
                            else: dia = str(dia)
                            cmp = now().strftime('%m-%Y')
                            data = f"{dia}-{cmp}"
                            wb_final = self.create_rl_vendas(data, "DD-MM-YYYY", cr, arquivos, opr)

                    case 'mes':
                        mes = int(valorF)
                        if mes < 10: mes = f"0{mes}"
                        else: mes = str(mes)
                        ano = now().strftime('%Y')
                        data = f"{mes}-{ano}"
                        wb_final = self.create_rl_vendas(data, 'MM-YYYY', cr, arquivos, opr)

                    case 'ano':
                        data = str(valorF)
                        if len(data) == 4: # Confere se o ano está correta
                            wb_final = self.create_rl_vendas(data, 'YYYY', cr, arquivos, opr)
                        else: return jsonify('Ano invalido, Formato YYYY requirido'), 401
                        
                for arquivo in arquivos:
                    wb = load_workbook(arquivo)
                    for nome in wb.sheetnames:
                        ws = wb[nome]
                        nova_sheet = wb_final.create_sheet(title=arquivo.replace('.xlsx', '').replace('src/', '').capitalize())

                    for row in ws.iter_rows(values_only=True):
                        nova_sheet.append(row)
                    remove(arquivo)
                filename = f'src/Vendas_{data}_{cr}_{now().strftime('%d%m%Y %H%M')}.xlsx'
                wb_final.save(filename)
                return jsonify(filename), 200
            return jsonify('Mes obrigatorio'), 400
        return jsonify('Credenciais Invalidas'), 401

    # Lojas ===================================================
    def get_dados_loja(self, cr):
        if cr:
            res = self.lcons("select cpf_cnpj, nome_loja, bairro, cep, cidade, estado, rua, sistema, telefone, email from lojas where cr = '%s'",  cr)
            cnpj, nome, bairro, cep, cidade, estado, rua, sistema, telefone, email = res
            loja = {
                'nome': nome,
                'bairro': bairro,
                'cep': cep,
                'cidade': cidade,
                'estado': estado,
                'rua': rua,
                'sistema': sistema,
                'telefone': telefone,
                'email': email,
            }
            if len(cnpj) == 14: loja['cnpj'] = cnpj
            else: loja['cpf'] = cnpj
            logo = self.cons("select logo from config where cr = '%s'", cr)[0]
            loja['logo'] = logo
            
            return jsonify(loja), 200
        return jsonify('Credenciais Invalidas'), 401